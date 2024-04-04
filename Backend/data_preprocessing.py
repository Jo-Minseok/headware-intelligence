import openpyxl

pre_file = openpyxl.load_workbook('original data.xlsx')
pre_sheet = pre_file.active

processing_dict = {}
for row in pre_sheet.iter_rows(min_row=2, values_only = True):
    if not row[3]:
        continue
    key = row[2] + ' ' + row[3]
    if key in processing_dict:
        continue
    processing_dict[key] = [row[5], row[6]]

processing_data = []
for k, v in processing_dict.items():
    processing_data.append([k, v[0], v[1]])

processing_file = openpyxl.Workbook()
processing_sheet = processing_file.active

for i in range(len(processing_data)):
    for j, k in zip('ABC', range(3)):
        processing_sheet[j + str(i + 1)] = processing_data[i][k]

processing_file.save('processing data.xlsx')