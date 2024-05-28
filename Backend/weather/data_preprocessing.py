import openpyxl

pre_sheet = openpyxl.load_workbook('./weather/original data.xlsx').active

processing_dict = {}
for row in pre_sheet.iter_rows(min_row=2, values_only = True):
    if not row[3]:
        continue
    key = row[2] + ' ' + row[3]
    if key in processing_dict:
        continue
    processing_dict[key] = [row[5], row[6], row[14], row[13]]

processing_data = [[k, v[0], v[1], v[2], v[3]] for k, v in processing_dict.items()]

processing_file = openpyxl.Workbook()
processing_sheet = processing_file.active

for i in range(len(processing_data)):
    for j, k in zip('ABCDE', range(5)):
        processing_sheet[j + str(i + 1)] = processing_data[i][k]

processing_file.save('./weather/processing data.xlsx')
